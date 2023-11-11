import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook

class MantenimientoVehiculosApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Mantenimiento de Vehículos")

        # Crear la interfaz de usuario
        self.codigo_label = ttk.Label(root, text="Código:")
        self.codigo_entry = ttk.Entry(root)
        self.marca_label = ttk.Label(root, text="Marca:")
        self.marca_entry = ttk.Entry(root)
        self.modelo_label = ttk.Label(root, text="Modelo:")
        self.modelo_entry = ttk.Entry(root)
        self.precio_label = ttk.Label(root, text="Precio:")
        self.precio_entry = ttk.Entry(root)
        self.kilometraje_label = ttk.Label(root, text="Kilometraje:")
        self.kilometraje_entry = ttk.Entry(root)

        self.guardar_button = ttk.Button(root, text="Guardar", command=self.guardar_vehiculo)
        self.editar_button = ttk.Button(root, text="Editar", command=self.editar_vehiculo)
        self.eliminar_button = ttk.Button(root, text="Eliminar", command=self.eliminar_vehiculo)
        self.listar_button = ttk.Button(root, text="Listar", command=self.listar_vehiculos)

        # Posicionar elementos en la interfaz
        self.codigo_label.grid(row=0, column=0, padx=5, pady=5)
        self.codigo_entry.grid(row=0, column=1, padx=5, pady=5)
        self.marca_label.grid(row=1, column=0, padx=5, pady=5)
        self.marca_entry.grid(row=1, column=1, padx=5, pady=5)
        self.modelo_label.grid(row=2, column=0, padx=5, pady=5)
        self.modelo_entry.grid(row=2, column=1, padx=5, pady=5)
        self.precio_label.grid(row=3, column=0, padx=5, pady=5)
        self.precio_entry.grid(row=3, column=1, padx=5, pady=5)
        self.kilometraje_label.grid(row=4, column=0, padx=5, pady=5)
        self.kilometraje_entry.grid(row=4, column=1, padx=5, pady=5)

        self.guardar_button.grid(row=5, column=0, columnspan=2, pady=10)
        self.editar_button.grid(row=6, column=0, columnspan=2, pady=10)
        self.eliminar_button.grid(row=7, column=0, columnspan=2, pady=10)
        self.listar_button.grid(row=8, column=0, columnspan=2, pady=10)

    def guardar_vehiculo(self):
        codigo = self.codigo_entry.get()
        marca = self.marca_entry.get()
        modelo = self.modelo_entry.get()
        precio = float(self.precio_entry.get())
        kilometraje = int(self.kilometraje_entry.get())

        vehiculo = [codigo, marca, modelo, precio, kilometraje]

        # Guardar vehículo en el archivo Excel
        self.guardar_en_excel(vehiculo)

        # Limpiar los campos después de guardar
        self.limpiar_campos()

    def guardar_en_excel(self, vehiculo):
        workbook = load_workbook("vehiculos.xlsx")
        sheet = workbook["listado"]

        # Obtener la última fila para agregar el nuevo vehículo
        last_row = sheet.max_row + 1

        for col, value in enumerate(vehiculo, start=1):
            sheet.cell(row=last_row, column=col, value=value)

        workbook.save("vehiculos.xlsx")

    def editar_vehiculo(self):
        # Implementar la funcionalidad de editar un vehículo
        pass

    def eliminar_vehiculo(self):
        # Implementar la funcionalidad de eliminar un vehículo
        pass

    def listar_vehiculos(self):
        # Implementar la funcionalidad de listar vehículos
        pass

    def limpiar_campos(self):
        self.codigo_entry.delete(0, tk.END)
        self.marca_entry.delete(0, tk.END)
        self.modelo_entry.delete(0, tk.END)
        self.precio_entry.delete(0, tk.END)
        self.kilometraje_entry.delete(0, tk.END)

if __name__ == "__main__":
    root = tk.Tk()
    app = MantenimientoVehiculosApp(root)
    root.mainloop()           
                     